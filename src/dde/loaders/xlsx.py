"""Guarded XLSX loader with explicit formula and hidden-sheet evidence."""

from __future__ import annotations

import json
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile

from defusedxml import ElementTree
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from dde.config import Settings
from dde.errors import CorruptInputError, InputError, InputLimitError, UnsupportedInputError
from dde.loaders.base import LoadedDocument, digest, safe_filename
from dde.models import LoaderNotice, LoaderNoticeCode, Severity

_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CANONICAL_PREFIX = (
    "XLSX visible sheets as JSON rows (formula text and cached values are labeled separately):\n"
)
_REQUIRED_PARTS = frozenset({"[Content_Types].xml", "xl/workbook.xml"})
_ACTIVE_PART_MARKERS = (
    "vbaproject.bin",
    "xl/activex/",
    "xl/embeddings/",
    "xl/macrosheets/",
    "xl/dialogsheets/",
    "xl/connections.xml",
    "customui/",
)
_ACTIVE_CONTENT_TYPE_MARKERS = (
    "macroenabled",
    "vbaproject",
    "activex",
    "oleobject",
    "macrosheet",
    "externallink",
    "connections",
    "customui",
)
_ACTIVE_RELATIONSHIP_NAMES = frozenset(
    {
        "vbaproject",
        "activexcontrol",
        "activexcontrolbinary",
        "oleobject",
        "package",
        "macrosheet",
        "xlmacrosheet",
        "intlmacrosheet",
        "externallink",
        "connections",
        "customui",
        "attachedtemplate",
    }
)


def _safe_part_name(name: str) -> bool:
    path = PurePosixPath(name.replace("\\", "/"))
    return not path.is_absolute() and ".." not in path.parts


def _preflight_archive(data: bytes, settings: Settings) -> None:
    try:
        with ZipFile(BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > settings.max_xlsx_zip_entries:
                raise InputLimitError(
                    f"XLSX has {len(entries)} ZIP entries; limit is {settings.max_xlsx_zip_entries}"
                )
            names = [entry.filename for entry in entries]
            if len(names) != len(set(names)):
                raise CorruptInputError("XLSX contains duplicate ZIP part names")
            if any(not _safe_part_name(name) for name in names):
                raise CorruptInputError("XLSX contains an unsafe ZIP part name")
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise UnsupportedInputError("Encrypted XLSX ZIP entries are not supported")
            uncompressed_bytes = sum(entry.file_size for entry in entries)
            if uncompressed_bytes > settings.max_xlsx_uncompressed_bytes:
                raise InputLimitError(
                    f"XLSX declares {uncompressed_bytes} uncompressed bytes; "
                    f"limit is {settings.max_xlsx_uncompressed_bytes}"
                )
            name_set = set(names)
            missing = sorted(_REQUIRED_PARTS - name_set)
            if missing:
                raise CorruptInputError(
                    "XLSX is missing required package parts: " + ", ".join(missing)
                )
            content_types_root = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            for content_type in content_types_root:
                declared_type = content_type.attrib.get("ContentType", "").casefold()
                if any(marker in declared_type for marker in _ACTIVE_CONTENT_TYPE_MARKERS):
                    raise UnsupportedInputError("XLSX active content is not supported")
            workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            declared_sheets = sum(
                1 for element in workbook_root.iter() if element.tag.endswith("}sheet")
            )
            if declared_sheets > settings.max_sheets:
                raise InputLimitError(
                    f"XLSX has {declared_sheets} sheets; limit is {settings.max_sheets}"
                )
            lower_names = [name.casefold() for name in names]
            if any(marker in name for name in lower_names for marker in _ACTIVE_PART_MARKERS):
                raise UnsupportedInputError("XLSX active content is not supported")
            for name in names:
                if not name.casefold().endswith(".rels"):
                    continue
                root = ElementTree.fromstring(archive.read(name))
                for relationship in root:
                    if relationship.attrib.get("TargetMode", "").casefold() == "external":
                        raise UnsupportedInputError("XLSX external relationships are not supported")
                    relationship_type = relationship.attrib.get("Type", "").casefold()
                    relationship_name = relationship_type.rstrip("/").rsplit("/", 1)[-1]
                    if relationship_name in _ACTIVE_RELATIONSHIP_NAMES:
                        raise UnsupportedInputError("XLSX active content is not supported")
            bad_part = archive.testzip()
            if bad_part is not None:
                raise CorruptInputError(f"XLSX ZIP checksum failed for {bad_part}")
    except (BadZipFile, DefusedXmlException, ElementTree.ParseError) as exc:
        raise CorruptInputError("XLSX package is corrupt") from exc


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float):
        return repr(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _bounded_cell_text(
    value: Any, settings: Settings, sheet_index: int, row: int, column: int
) -> str | None:
    text = _cell_text(value)
    if text is not None and len(text) > settings.max_cell_chars:
        raise InputLimitError(
            f"XLSX cell {sheet_index}:{row},{column} has {len(text)} characters; "
            f"limit is {settings.max_cell_chars}"
        )
    return text


def _canonical_cell(
    formula_cell: Cell | MergedCell,
    cached_cell: Cell | MergedCell,
    settings: Settings,
    sheet_index: int,
    row: int,
    column: int,
) -> tuple[object, bool, bool]:
    if formula_cell.data_type == "f":
        formula = _bounded_cell_text(formula_cell.value, settings, sheet_index, row, column)
        cached = _bounded_cell_text(cached_cell.value, settings, sheet_index, row, column)
        return {"formula": formula, "cached_value": cached}, True, cached is None
    value = _bounded_cell_text(formula_cell.value, settings, sheet_index, row, column)
    if formula_cell.data_type == "e" and value is not None:
        return {"error": value}, False, False
    return value, False, False


def _visible_sheets(workbook: Any) -> list[Worksheet]:
    return [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]


def load_xlsx(path: Path, data: bytes, settings: Settings) -> LoadedDocument:
    _preflight_archive(data, settings)
    formula_workbook = None
    cached_workbook = None
    try:
        formula_workbook = load_workbook(
            BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
        cached_workbook = load_workbook(
            BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
        all_sheets = formula_workbook.worksheets
        if len(all_sheets) > settings.max_sheets:
            raise InputLimitError(
                f"XLSX has {len(all_sheets)} sheets; limit is {settings.max_sheets}"
            )
        visible_formula = _visible_sheets(formula_workbook)
        visible_cached = _visible_sheets(cached_workbook)
        if not visible_formula:
            raise InputError("XLSX contains no visible worksheets")
        if [sheet.title for sheet in visible_formula] != [sheet.title for sheet in visible_cached]:
            raise CorruptInputError("XLSX formula and cached sheet views do not align")

        notices: list[LoaderNotice] = []
        hidden_count = len(all_sheets) - len(visible_formula)
        if hidden_count:
            notices.append(
                LoaderNotice(
                    code=LoaderNoticeCode.XLSX_HIDDEN_SHEET_SKIPPED,
                    severity=Severity.WARNING,
                    message=f"{hidden_count} hidden XLSX sheet(s) were skipped",
                    field=None,
                )
            )

        parts = [_CANONICAL_PREFIX]
        canonical_chars = len(_CANONICAL_PREFIX)
        total_rows = 0
        formula_present = False
        cache_missing = False
        for sheet_index, (formula_sheet, cached_sheet) in enumerate(
            zip(visible_formula, visible_cached, strict=True), start=1
        ):
            max_row = formula_sheet.max_row or 0
            max_column = formula_sheet.max_column or 0
            if max_column > settings.max_tabular_columns:
                raise InputLimitError(
                    f"XLSX sheet {sheet_index} has {max_column} columns; "
                    f"limit is {settings.max_tabular_columns}"
                )
            total_rows += max_row
            if total_rows > settings.max_tabular_rows:
                raise InputLimitError(f"XLSX has more than {settings.max_tabular_rows} total rows")
            heading = (
                f"sheet {sheet_index}: {json.dumps(formula_sheet.title, ensure_ascii=False)}\n"
            )
            canonical_chars += len(heading)
            if canonical_chars > settings.max_tabular_chars:
                raise InputLimitError(
                    f"Canonical XLSX text exceeds {settings.max_tabular_chars} characters"
                )
            parts.append(heading)
            formula_rows = formula_sheet.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_column
            )
            cached_rows = cached_sheet.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_column
            )
            for row_index, (formula_row, cached_row) in enumerate(
                zip(formula_rows, cached_rows, strict=True), start=1
            ):
                values: list[object] = []
                for column, (formula_cell, cached_cell) in enumerate(
                    zip(formula_row, cached_row, strict=True), start=1
                ):
                    value, is_formula, is_missing = _canonical_cell(
                        formula_cell,
                        cached_cell,
                        settings,
                        sheet_index,
                        row_index,
                        column,
                    )
                    values.append(value)
                    formula_present |= is_formula
                    cache_missing |= is_missing
                while values and values[-1] is None:
                    values.pop()
                line = f"row {row_index}: {json.dumps(values, ensure_ascii=False)}\n"
                canonical_chars += len(line)
                if canonical_chars > settings.max_tabular_chars:
                    raise InputLimitError(
                        f"Canonical XLSX text exceeds {settings.max_tabular_chars} characters"
                    )
                parts.append(line)

        if formula_present:
            notices.append(
                LoaderNotice(
                    code=LoaderNoticeCode.XLSX_FORMULA_PRESENT,
                    severity=Severity.WARNING,
                    message="XLSX formulas are preserved as text and are not executed",
                    field=None,
                )
            )
        if cache_missing:
            notices.append(
                LoaderNotice(
                    code=LoaderNoticeCode.XLSX_FORMULA_CACHE_MISSING,
                    severity=Severity.WARNING,
                    message="One or more XLSX formulas have no cached value",
                    field=None,
                )
            )

        return LoadedDocument(
            file_name=safe_filename(path),
            media_type=_MEDIA_TYPE,
            sha256=digest(data),
            byte_count=len(data),
            page_count=None,
            sheet_count=len(all_sheets),
            text="".join(parts),
            images=(),
            notices=tuple(notices),
        )
    except (InvalidFileException, ValueError, KeyError, IndexError, TypeError) as exc:
        raise CorruptInputError("XLSX workbook structure is corrupt") from exc
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
