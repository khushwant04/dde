from __future__ import annotations

from collections.abc import Callable
from datetime import date
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from dde.config import Settings
from dde.errors import CorruptInputError, InputError, InputLimitError, UnsupportedInputError
from dde.loaders import load_document
from dde.models import LoaderNoticeCode, Severity


def settings(**overrides: object) -> Settings:
    return Settings(_env_file=None, **overrides)


def workbook_bytes(configure: Callable[[Workbook], None] | None = None) -> bytes:
    workbook = Workbook()
    if configure is not None:
        configure(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def write_xlsx(tmp_path: Path, data: bytes, name: str = "document.xlsx") -> Path:
    path = tmp_path / name
    path.write_bytes(data)
    return path


def rewrite_zip(
    data: bytes,
    transform: Callable[[str, bytes], bytes] | None = None,
    additions: dict[str, bytes] | None = None,
) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(data)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            content = source.read(info.filename)
            if transform is not None:
                content = transform(info.filename, content)
            target.writestr(info.filename, content)
        for name, content in (additions or {}).items():
            target.writestr(name, content)
    return output.getvalue()


def patch_central_directory(data: bytes, member: str, field_offset: int, value: bytes) -> bytes:
    raw = bytearray(data)
    position = 0
    while True:
        position = raw.find(b"PK\x01\x02", position)
        if position < 0:
            raise AssertionError(f"central directory member not found: {member}")
        name_length = int.from_bytes(raw[position + 28 : position + 30], "little")
        extra_length = int.from_bytes(raw[position + 30 : position + 32], "little")
        comment_length = int.from_bytes(raw[position + 32 : position + 34], "little")
        name_start = position + 46
        name = bytes(raw[name_start : name_start + name_length]).decode()
        if name == member:
            start = position + field_offset
            raw[start : start + len(value)] = value
            return bytes(raw)
        position = name_start + name_length + extra_length + comment_length


def test_xlsx_canonicalizes_visible_sheets_and_scalar_types(tmp_path: Path) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Invoice"
        sheet.append(["ID", "Amount", "Issued", "Paid"])
        sheet.append(["INV-1", 12.5, date(2026, 8, 13), True])
        second = workbook.create_sheet("Lines")
        second.append(["Widget", 2, 6.25])

    loaded = load_document(write_xlsx(tmp_path, workbook_bytes(configure)), settings())
    assert loaded.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert loaded.page_count is None
    assert loaded.sheet_count == 2
    assert loaded.notices == ()
    assert loaded.text is not None
    assert 'sheet 1: "Invoice"' in loaded.text
    assert '["INV-1", "12.5", "2026-08-13T00:00:00", "true"]' in loaded.text
    assert 'sheet 2: "Lines"' in loaded.text


def test_xlsx_skips_hidden_sheets_with_review_notice(tmp_path: Path) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active.append(["visible"])
        hidden = workbook.create_sheet("Internal")
        hidden.append(["secret formula input"])
        hidden.sheet_state = "hidden"

    loaded = load_document(write_xlsx(tmp_path, workbook_bytes(configure)), settings())
    assert loaded.sheet_count == 2
    assert loaded.text is not None and "secret formula input" not in loaded.text
    assert [notice.code for notice in loaded.notices] == [
        LoaderNoticeCode.XLSX_HIDDEN_SHEET_SKIPPED
    ]
    assert loaded.notices[0].severity == Severity.WARNING


def test_xlsx_formula_text_is_preserved_but_never_executed(tmp_path: Path) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.append([1, 2, "=SUM(A1:B1)"])

    loaded = load_document(write_xlsx(tmp_path, workbook_bytes(configure)), settings())
    assert loaded.text is not None
    assert '{"formula": "=SUM(A1:B1)", "cached_value": null}' in loaded.text
    assert [notice.code for notice in loaded.notices] == [
        LoaderNoticeCode.XLSX_FORMULA_PRESENT,
        LoaderNoticeCode.XLSX_FORMULA_CACHE_MISSING,
    ]
    assert all(notice.severity == Severity.WARNING for notice in loaded.notices)


def test_xlsx_cached_formula_value_is_labeled_separately(tmp_path: Path) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active.append([1, 2, "=SUM(A1:B1)"])

    def add_cache(name: str, content: bytes) -> bytes:
        if name.startswith("xl/worksheets/sheet"):
            return content.replace(b"<f>SUM(A1:B1)</f><v />", b"<f>SUM(A1:B1)</f><v>3</v>")
        return content

    data = rewrite_zip(workbook_bytes(configure), transform=add_cache)
    loaded = load_document(write_xlsx(tmp_path, data), settings())
    assert loaded.text is not None
    assert '{"formula": "=SUM(A1:B1)", "cached_value": "3"}' in loaded.text
    assert [notice.code for notice in loaded.notices] == [LoaderNoticeCode.XLSX_FORMULA_PRESENT]


def test_xlsx_rejects_external_relationships(tmp_path: Path) -> None:
    def add_external(name: str, content: bytes) -> bytes:
        if name == "xl/_rels/workbook.xml.rels":
            relationship = (
                b'<Relationship Id="rIdExternal" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/externalLink" Target="https://example.invalid/book.xlsx" '
                b'TargetMode="External"/>'
            )
            return content.replace(b"</Relationships>", relationship + b"</Relationships>")
        return content

    data = rewrite_zip(workbook_bytes(), transform=add_external)
    with pytest.raises(UnsupportedInputError, match="external relationships"):
        load_document(write_xlsx(tmp_path, data), settings())


def test_xlsx_rejects_active_content(tmp_path: Path) -> None:
    data = rewrite_zip(
        workbook_bytes(), additions={"xl/vbaProject.bin": b"not executable test data"}
    )
    with pytest.raises(UnsupportedInputError, match="active content"):
        load_document(write_xlsx(tmp_path, data), settings())


def test_xlsx_rejects_renamed_active_content_by_semantic_types(tmp_path: Path) -> None:
    def declare_active_content(name: str, content: bytes) -> bytes:
        if name == "[Content_Types].xml":
            override = (
                b'<Override PartName="/xl/payload.bin" '
                b'ContentType="application/vnd.ms-office.vbaProject"/>'
            )
            return content.replace(b"</Types>", override + b"</Types>")
        if name == "xl/_rels/workbook.xml.rels":
            relationship = (
                b'<Relationship Id="rIdPayload" '
                b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                b'Target="payload.bin"/>'
            )
            return content.replace(b"</Relationships>", relationship + b"</Relationships>")
        return content

    data = rewrite_zip(
        workbook_bytes(),
        transform=declare_active_content,
        additions={"xl/payload.bin": b"renamed inert test payload"},
    )
    with pytest.raises(UnsupportedInputError, match="active content"):
        load_document(write_xlsx(tmp_path, data), settings())


@pytest.mark.parametrize(
    ("overrides", "message", "configure"),
    [
        (
            {"DDE_MAX_SHEETS": 1},
            "sheets",
            lambda workbook: workbook.create_sheet("Second"),
        ),
        (
            {"DDE_MAX_TABULAR_ROWS": 1},
            "total rows",
            lambda workbook: (workbook.active.append(["a"]), workbook.active.append(["b"])),
        ),
        (
            {"DDE_MAX_TABULAR_COLUMNS": 1},
            "columns",
            lambda workbook: workbook.active.append(["a", "b"]),
        ),
        (
            {"DDE_MAX_CELL_CHARS": 3},
            "characters",
            lambda workbook: workbook.active.append(["abcd"]),
        ),
        (
            {"DDE_MAX_TABULAR_CHARS": 20},
            "Canonical XLSX text",
            lambda workbook: workbook.active.append(["a"]),
        ),
    ],
)
def test_xlsx_enforces_workbook_and_tabular_limits(
    tmp_path: Path,
    overrides: dict[str, int],
    message: str,
    configure: Callable[[Workbook], object],
) -> None:
    with pytest.raises(InputLimitError, match=message):
        load_document(write_xlsx(tmp_path, workbook_bytes(configure)), settings(**overrides))


@pytest.mark.parametrize(
    ("overrides", "message"),
    [
        ({"DDE_MAX_XLSX_ZIP_ENTRIES": 1}, "ZIP entries"),
        ({"DDE_MAX_XLSX_UNCOMPRESSED_BYTES": 1}, "uncompressed bytes"),
    ],
)
def test_xlsx_enforces_archive_limits(
    tmp_path: Path, overrides: dict[str, int], message: str
) -> None:
    with pytest.raises(InputLimitError, match=message):
        load_document(write_xlsx(tmp_path, workbook_bytes()), settings(**overrides))


def test_xlsx_rejects_duplicate_and_unsafe_part_names(tmp_path: Path) -> None:
    original = workbook_bytes()
    output = BytesIO()
    with (
        pytest.warns(UserWarning, match="Duplicate name"),
        ZipFile(BytesIO(original)) as source,
        ZipFile(output, "w", ZIP_DEFLATED) as target,
    ):
        for info in source.infolist():
            target.writestr(info.filename, source.read(info.filename))
        duplicate = source.infolist()[0]
        target.writestr(duplicate.filename, source.read(duplicate.filename))
    with pytest.raises(CorruptInputError, match="duplicate ZIP part names"):
        load_document(write_xlsx(tmp_path, output.getvalue()), settings())

    unsafe = rewrite_zip(original, additions={"../payload.bin": b"test"})
    with pytest.raises(CorruptInputError, match="unsafe ZIP part name"):
        load_document(write_xlsx(tmp_path, unsafe), settings())


def test_xlsx_rejects_encrypted_entries_and_missing_parts(tmp_path: Path) -> None:
    encrypted = patch_central_directory(
        workbook_bytes(), "[Content_Types].xml", 8, (1).to_bytes(2, "little")
    )
    with pytest.raises(UnsupportedInputError, match="Encrypted XLSX"):
        load_document(write_xlsx(tmp_path, encrypted), settings())

    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr("placeholder.txt", "not a workbook")
    with pytest.raises(CorruptInputError, match="missing required package parts"):
        load_document(write_xlsx(tmp_path, output.getvalue()), settings())


def test_xlsx_rejects_bad_checksums_and_no_visible_sheets(tmp_path: Path) -> None:
    bad_checksum = patch_central_directory(workbook_bytes(), "docProps/app.xml", 16, bytes(4))
    with pytest.raises(CorruptInputError, match="ZIP checksum failed"):
        load_document(write_xlsx(tmp_path, bad_checksum), settings())

    def hide_only_sheet(name: str, content: bytes) -> bytes:
        if name == "xl/workbook.xml":
            return content.replace(b'state="visible"', b'state="hidden"', 1)
        return content

    hidden = rewrite_zip(workbook_bytes(), transform=hide_only_sheet)
    with pytest.raises(InputError, match="no visible worksheets"):
        load_document(write_xlsx(tmp_path, hidden), settings())


def test_xlsx_rejects_corrupt_or_wrong_extension_packages(tmp_path: Path) -> None:
    corrupt = write_xlsx(tmp_path, b"PK\x03\x04not-a-workbook")
    with pytest.raises(CorruptInputError):
        load_document(corrupt, settings())
    wrong_extension = write_xlsx(tmp_path, workbook_bytes(), "document.zip")
    with pytest.raises(UnsupportedInputError, match=r"requires a \.xlsx extension"):
        load_document(wrong_extension, settings())
